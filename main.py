from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os
from datetime import datetime
import glob

def listar_arquivos_pdf(origem):
    """Lista os arquivos PDF em um diretório."""
    arquivos_pdf = glob.glob(os.path.join(origem, '*.pdf'))
    nomes_arquivos_pdf = [os.path.basename(arquivo) for arquivo in arquivos_pdf]
    return sorted(nomes_arquivos_pdf)

def salvar_em_excel(nomes_arquivos, nome_arquivo_excel):
    """Salva os nomes dos arquivos em um arquivo Excel no mesmo diretório do script."""
    wb = Workbook()
    ws = wb.active

    # Definindo estilos para o cabeçalho
    header_font = Font(size=16, bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    header_border = Border(left=Side(style='thin'),
                           right=Side(style='thin'),
                           top=Side(style='thin'),
                           bottom=Side(style='thin'))

    # Definindo estilos para os nomes dos processos
    processo_font = Font(size=12)
    processo_alignment = Alignment(horizontal="left", vertical="center")
    processo_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

    # Adicionando o cabeçalho
    ws.append(['Relatório de Processos Cadastrados'])

    # Aplicando estilos ao cabeçalho
    header_cell = ws['A1']
    header_cell.font = header_font
    header_cell.alignment = header_alignment
    header_cell.fill = header_fill
    header_cell.border = header_border

    # Adicionando os nomes dos arquivos (processos)
    for nome in nomes_arquivos:
        # Removendo a extensão .pdf dos nomes dos arquivos
        nome_sem_extensao = os.path.splitext(nome)[0]
        ws.append([nome_sem_extensao])

    # Adicionando a linha final com data e hora atuais, quantidade de processos e assinatura
    data_atual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    quantidade_processos = len(nomes_arquivos)
    assinatura = (
        "Este relatório foi gerado automaticamente e contém todos os processos cadastrados até a data e hora especificadas."
    )
    ws.append([])
    ws.append([f"Relatório gerado em: {data_atual}"])
    ws.append([f"Quantidade de processos: {quantidade_processos}"])
    ws.append([assinatura])

    # Aplicando estilos às células finais
    for row in ws.iter_rows(min_row=ws.max_row-2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(italic=True)
            cell.alignment = Alignment(horizontal="left")
            cell.border = header_border

    # Aplicando estilos aos nomes dos processos
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row-4, min_col=1, max_col=1):
        for cell in row:
            cell.font = processo_font
            cell.alignment = processo_alignment
            cell.border = processo_border

    # Ajustar a largura da coluna para caber o conteúdo
    ws.column_dimensions['A'].width = 50

    # Salvar o arquivo Excel no mesmo diretório do script
    wb.save(nome_arquivo_excel)

    # Mensagem de sucesso
    print("Processos cadastrados salvos com sucesso em", nome_arquivo_excel)

if __name__ == "__main__":
    # Diretório atual onde o script está sendo executado
    diretorio_atual = os.path.dirname(os.path.abspath(__file__))

    # Nome do arquivo Excel de saída (no mesmo diretório do script)
    nome_arquivo_excel = os.path.join(diretorio_atual, "Relatorio_de_Processos.xlsx")

    # Lista de arquivos PDF no diretório atual
    nomes_arquivos_pdf = listar_arquivos_pdf(diretorio_atual)

    # Salvar os nomes dos arquivos em um arquivo Excel no mesmo diretório do script
    salvar_em_excel(nomes_arquivos_pdf, nome_arquivo_excel)
